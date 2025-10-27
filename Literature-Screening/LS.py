import os
import time
import math
import requests
import pandas as pd
import xml.etree.ElementTree as ET
from datetime import datetime
import tkinter as tk
from tkinter import messagebox

BASE_URL = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/"
TOOL = "pubmed_automation"
EMAIL = os.getenv("NCBI_EMAIL", "sales@iziel.com")  # set env var NCBI_EMAIL
API_KEY = os.getenv("NCBI_API_KEY")  # optional but recommended

# ---------- UI ----------
def collect_inputs():
    # if user chooses date filter
    if date_filter_var.get():
        from_date = from_date_entry.get().strip()
        to_date = to_date_entry.get().strip()

        # validate only if user entered a value
        if from_date:
            try:
                datetime.strptime(from_date, "%Y-%m-%d")
            except ValueError:
                messagebox.showerror("Input Error", "FROM date must be in YYYY-MM-DD format.")
                return

        if to_date:
            try:
                datetime.strptime(to_date, "%Y-%m-%d")
            except ValueError:
                messagebox.showerror("Input Error", "TO date must be in YYYY-MM-DD format.")
                return

        root.from_date = from_date if from_date else None
        root.to_date = to_date if to_date else None
    else:
        root.from_date = None
        root.to_date = None


    root.apply_abstract = abstract_var.get()
    root.apply_free_full_text = free_full_text_var.get()
    root.apply_full_text = full_text_var.get()
    root.destroy()

root = tk.Tk()
root.title("PubMed Filter Selection")

# NEW checkbox to let user decide
date_filter_var = tk.BooleanVar()
tk.Checkbutton(root, text="Apply Date Filter", variable=date_filter_var).grid(row=0, column=0, sticky="w", columnspan=2)

tk.Label(root, text="Enter FROM date (YYYY-MM-DD):").grid(row=1, column=0, sticky="w")
from_date_entry = tk.Entry(root, width=20)
from_date_entry.grid(row=1, column=1, padx=5, pady=5)

tk.Label(root, text="Enter TO date (YYYY-MM-DD):").grid(row=2, column=0, sticky="w")
to_date_entry = tk.Entry(root, width=20)
to_date_entry.grid(row=2, column=1, padx=5, pady=5)

tk.Label(root, text="Text Availability Filters:").grid(row=3, column=0, sticky="w", pady=(10, 0))
abstract_var = tk.BooleanVar()
free_full_text_var = tk.BooleanVar()
full_text_var = tk.BooleanVar()
tk.Checkbutton(root, text="Abstract", variable=abstract_var).grid(row=4, column=0, sticky="w")
tk.Checkbutton(root, text="Free full text", variable=free_full_text_var).grid(row=5, column=0, sticky="w")
tk.Checkbutton(root, text="Full text", variable=full_text_var).grid(row=6, column=0, sticky="w")

tk.Button(root, text="Submit", command=collect_inputs).grid(row=7, column=0, columnspan=2, pady=10)

#  Pre-fill default dates for testing
from_date_entry.insert(0, "2024-08-02")
to_date_entry.insert(0, "2025-07-30")

root.mainloop()


FROM_DATE = root.from_date
TO_DATE = root.to_date
APPLY_ABSTRACT = root.apply_abstract
APPLY_FREE = root.apply_free_full_text
APPLY_FULL = root.apply_full_text

# ---------- Paths ----------
script_dir = os.path.dirname(os.path.abspath(__file__))
csv_dir = os.path.join(script_dir, "All-CSV")
os.makedirs(csv_dir, exist_ok=True)

# ---------- Load keywords ----------
df = pd.read_excel("keywords.xlsx", header=0)
df.columns = df.columns.str.strip()

if "Keyword No." not in df.columns or "Keywords" not in df.columns:
    raise ValueError("Required columns 'Keyword No.' and 'Keywords' not found in Excel.")

filtered_df = df[df["Keyword No."].astype(str).str.startswith("#")].copy()
filters_col = filtered_df["Filters"] if "Filters" in filtered_df.columns else pd.Series([""] * len(filtered_df))

# Use "Keywords" instead of "Keyword"
keywords = list(zip(filtered_df["Keyword No."], filtered_df["Keywords"], filters_col.fillna("")))

print(f"✅ Filtered {len(keywords)} keywords starting with '#'")

# ---------- Helpers ----------
session = requests.Session()
DEFAULT_SLEEP = 0.34 if not API_KEY else 0.12  # ~3 r/s without key, ~8-10 r/s with key

def _common_params():
    p = {"tool": TOOL, "email": EMAIL}
    if API_KEY:
        p["api_key"] = API_KEY
    return p

def build_query(keyword: str, filters_csv: str) -> str:
    """Compose the PubMed term with English+Humans + availability + article types."""
    parts = [f"({keyword})", "english[lang]", "humans[mh]"]

    # availability filters (OR logic, like PubMed UI)
    avail_terms = []
    if APPLY_ABSTRACT:
        avail_terms.append("hasabstract[text]")
    if APPLY_FREE:
        avail_terms.append("free full text[sb]")
    if APPLY_FULL:
        avail_terms.append("full text[sb]")
    if avail_terms:
        parts.append("(" + " OR ".join(avail_terms) + ")")

    # publication types (OR logic)
    if isinstance(filters_csv, str) and filters_csv.strip():
        types = [t.strip() for t in filters_csv.split(",") if t.strip()]
        if types:
            types_q = " OR ".join([f"\"{t}\"[Publication Type]" for t in types])
            parts.append(f"({types_q})")

    return " AND ".join(parts)


def esearch_with_history(term: str, mindate=None, maxdate=None):
    """Returns (count, query_key, webenv)."""
    params = {
        "db": "pubmed",
        "term": term,
        "retmode": "json",
        "usehistory": "y",
    }
    # only add date filters if provided
    # add date filters based on user input
    if mindate and maxdate:
        params.update({
            "datetype": "pdat",
            "mindate": mindate.replace("-", "/"),
            "maxdate": maxdate.replace("-", "/"),
        })
    elif mindate and not maxdate:
        params.update({
            "datetype": "pdat",
            "mindate": mindate.replace("-", "/"),
        })
    elif maxdate and not mindate:
        params.update({
            "datetype": "pdat",
            "maxdate": maxdate.replace("-", "/"),
        })

    params.update(_common_params())

    # 🔎 DEBUG: print the query & full URL
    debug_url = requests.Request("GET", BASE_URL + "esearch.fcgi", params=params).prepare().url
    print(f"   🔎 Final Query: {term}")
    if mindate and maxdate:
        print(f"   🔎 Date range: {mindate} → {maxdate}")
    # print(f"   🔎 Full URL: {debug_url}")

    r = session.get(BASE_URL + "esearch.fcgi", params=params, timeout=60)
    r.raise_for_status()
    data = r.json()
    res = data.get("esearchresult", {})
    count = int(res.get("count", 0))
    qk = res.get("querykey")
    we = res.get("webenv")
    return count, qk, we

def safe_text(el):
    return (el.text or "").strip() if el is not None else ""

def parse_date(article):
    y = article.findtext(".//Journal/JournalIssue/PubDate/Year")
    m = article.findtext(".//Journal/JournalIssue/PubDate/Month")
    d = article.findtext(".//Journal/JournalIssue/PubDate/Day")
    if y:
        parts = [y, m or "", d or ""]
        return "-".join([p for p in parts if p])
    md = article.findtext(".//Journal/JournalIssue/PubDate/MedlineDate")
    return (md or "").strip()

def parse_authors(article):
    names = []
    for a in article.findall(".//AuthorList/Author"):
        last = safe_text(a.find("LastName"))
        initials = safe_text(a.find("Initials"))
        if last or initials:
            names.append(f"{last} {initials}".strip())
    return ", ".join(names)

def parse_pubtypes(article):
    types = [safe_text(t) for t in article.findall(".//PublicationTypeList/PublicationType")]
    return ", ".join([t for t in types if t])

def parse_abstract(article):
    texts = []
    for at in article.findall(".//Abstract/AbstractText"):
        label = at.get("Label")
        content = (at.text or "").strip()
        if not content:
            continue
        if label:
            texts.append(f"{label}: {content}")
        else:
            texts.append(content)
    return "\n\n".join(texts)

def parse_doi(article):
    for aid in article.findall(".//ArticleIdList/ArticleId"):
        if aid.get("IdType") == "doi":
            return safe_text(aid)
    return ""

def efetch_batch(qk: str, we: str, retstart: int, retmax: int = 200):
    params = {
        "db": "pubmed",
        "query_key": qk,
        "webenv": we,
        "retstart": retstart,
        "retmax": retmax,
        "retmode": "xml",
    }
    params.update(_common_params())
    r = session.get(BASE_URL + "efetch.fcgi", params=params, timeout=120)
    r.raise_for_status()
    return r.text

def xml_to_records(xml_text: str):
    root = ET.fromstring(xml_text)
    records = []
    for art in root.findall(".//PubmedArticle"):
        pmid = safe_text(art.find(".//PMID"))
        title = safe_text(art.find(".//ArticleTitle"))
        journal = safe_text(art.find(".//Journal/Title"))
        pubdate = parse_date(art)
        authors = parse_authors(art)
        abstract = parse_abstract(art)
        pubtypes = parse_pubtypes(art)
        doi = parse_doi(art)
        records.append({
            "PMID": pmid,
            "Title": title,
            "Journal": journal,
            "PubDate": pubdate,
            "Authors": authors,
            "PublicationTypes": pubtypes,
            "DOI": doi,
            "Abstract": abstract,
            "PubMedURL": f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/" if pmid else ""
        })
    return records

# ---------- Main loop ----------
missed = []
hit_counts = {}   # NEW: store keyword hits

for keyword_no, keyword, filters_csv in keywords:
    print(f"\n🔍 Searching: {keyword} ({keyword_no})")
    term = build_query(keyword, filters_csv)
    try:
        count, qk, we = esearch_with_history(term, FROM_DATE, TO_DATE)
        print(f"   • Matches: {count}")

        hit_counts[keyword_no] = count   # NEW

        if count == 0:
            missed.append((keyword_no, keyword))
            continue

        all_rows = []
        BATCH = 200
        loops = math.ceil(count / BATCH)
        for i in range(loops):
            retstart = i * BATCH
            time.sleep(DEFAULT_SLEEP)
            xml_text = efetch_batch(qk, we, retstart, BATCH)
            rows = xml_to_records(xml_text)
            all_rows.extend(rows)
            print(f"   • Fetched {len(rows)} (total {len(all_rows)}/{count})")

        out_path = os.path.join(csv_dir, f"{keyword_no}.csv")
        pd.DataFrame(all_rows).to_csv(out_path, index=False)
        print(f"📁 Saved: {out_path}")

    except Exception as e:
        print(f"⚠️ Failed for {keyword} → {e}")
        missed.append((keyword_no, keyword))
        hit_counts[keyword_no] = 0   # NEW
        time.sleep(1.0)

# ---------- Missed ----------
if missed:
    pd.DataFrame(missed, columns=["Keyword No.", "Keywords"]).to_excel(
        os.path.join(script_dir, "Missed.xlsx"), index=False
    )
    print("⚠️ Missed keywords saved → Missed.xlsx")
else:
    print("✅ All keywords processed without misses.")

print(f"\nAll CSVs saved in: {csv_dir}")

# ---------- Update keywords.xlsx with hit counts ----------
if "Number of Hits" not in df.columns:
    df["Number of Hits"] = ""

for idx, row in df.iterrows():
    k_no = row["Keyword No."]
    if k_no in hit_counts:
        old_val = str(row.get("Number of Hits", "")).strip()
        new_val = str(hit_counts[k_no])
        if old_val and old_val.lower() != "nan":
            df.at[idx, "Number of Hits"] = f"{old_val}, {new_val}"
        else:
            df.at[idx, "Number of Hits"] = new_val

df.to_excel("keywords.xlsx", index=False)
print("✅ Updated keywords.xlsx with hit counts")

# Merge.py (fixed for API output)
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

script_dir = os.path.dirname(os.path.abspath(__file__))
csv_dir = os.path.join(script_dir, "All-CSV")
output_path = os.path.join(script_dir, "All-Merged.xlsx")

combined_rows = []
required_cols = ["PMID", "Title", "Journal", "PubDate", "Authors",
                 "PublicationTypes", "DOI", "Abstract", "PubMedURL"]

for file_name in sorted(os.listdir(csv_dir)):
    if file_name.endswith(".csv") and file_name.startswith("#"):
        keyword_no = os.path.splitext(file_name)[0]
        file_path = os.path.join(csv_dir, file_name)

        try:
            df = pd.read_csv(file_path)
            if not all(col in df.columns for col in required_cols):
                print(f"⚠️ Skipped {file_name}: Missing required columns.")
                continue

            for idx, row in df.iterrows():
                pmid = str(row["PMID"]).strip()
                combined_rows.append({
                    "KeywordNo": keyword_no,
                    "KeyCodeNo": f"{keyword_no}.{idx + 1}",
                    "PMID": pmid,
                    "Title": row["Title"],
                    "Journal": row["Journal"],
                    "PubDate": row["PubDate"],
                    "Authors": row["Authors"],
                    "PublicationTypes": row["PublicationTypes"],
                    "DOI": row["DOI"],
                    "Abstract": row["Abstract"],
                    "PubMedURL": row["PubMedURL"]
                })

        except Exception as e:
            print(f"❌ Error processing {file_name}: {e}")

if not combined_rows:
    print("⚠️ No data found in CSVs.")
    exit()

combined_df = pd.DataFrame(combined_rows)
combined_df.insert(0, "Sr.No", range(1, len(combined_df) + 1))

master_df = combined_df.drop_duplicates(subset="PMID", keep="first").reset_index(drop=True)
master_df["Sr.No"] = range(1, len(master_df) + 1)

duplicates = combined_df.duplicated(subset="PMID", keep=False)
highlight_df = combined_df.copy()

with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    combined_df.to_excel(writer, sheet_name="Combined", index=False)
    highlight_df.to_excel(writer, sheet_name="Duplicate", index=False)
    master_df.to_excel(writer, sheet_name="Master", index=False)

wb = load_workbook(output_path)
ws = wb["Duplicate"]
fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

pmid_col_index = list(highlight_df.columns).index("PMID") + 1
for row_idx, is_dup in enumerate(duplicates, start=2):
    if is_dup:
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col).fill = fill

wb.save(output_path)

print(f"✅ Excel file created with 3 sheets: {output_path}")
